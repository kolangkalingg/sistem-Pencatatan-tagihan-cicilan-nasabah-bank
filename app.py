import sqlite3
import openpyxl
from flask import Flask, render_template, request, redirect, session, flash, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import timedelta



# Konstanta ma file Excel
FILE_NAME = "data_tagihan.xlsx"

# Inisialisasi aplikasi Flask
app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Ganti dengan kunci rahasia yang aman
app.config['SESSION_PERMANENT'] = False  # Sesi tidak bersifat permanen


def init_sqlite_db():
    conn = sqlite3.connect("billing_system.db")
    cursor = conn.cursor()

    # Membuat tabel users jika belum ada
    cursor.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL UNIQUE,
        password TEXT NOT NULL,
        nama TEXT,
        alamat TEXT,
        no_hp TEXT,
        email TEXT,
        role TEXT DEFAULT 'user',  -- Tambahkan kolom role dengan default 'user'
        approved INTEGER DEFAULT 0 -- Tambahkan kolom approved dengan default 0 (belum disetujui)
    )''')

    # Membuat tabel obrolan jika belum ada
    cursor.execute('''CREATE TABLE IF NOT EXISTS obrolan (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL,
        message TEXT NOT NULL,
        timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    conn.commit()
    conn.close()
    add_default_admin()

def add_default_admin():
    conn = sqlite3.connect("billing_system.db")
    cursor = conn.cursor()

    # Periksa apakah admin sudah ada
    cursor.execute("SELECT * FROM users WHERE role = 'admin'")
    admin = cursor.fetchone()

    if not admin:
        hashed_password = generate_password_hash("admin123")  # Password default
        cursor.execute('''
            INSERT INTO users (username, password, nama, alamat, no_hp, email, role, approved)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', ("admin", hashed_password, "Admin", "Admin Address", "1234567890", "admin@example.com", "admin", 1))
        conn.commit()

    conn.close()

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session or session.get('role') != 'admin':
            flash("Hanya admin yang dapat mengakses halaman ini.")
            return redirect('/')
        return f(*args, **kwargs)
    return decorated_function



def initialize_excel():
    try:
        workbook = openpyxl.load_workbook(FILE_NAME)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    if "Data Cicilan yang Belum Dibayar" not in workbook.sheetnames:
        sheet = workbook.create_sheet("Data_Cicilan_yang_Belum_Dibayar")
        sheet.append([
            "No Kontrak", "Hari dan Tanggal", "Nama BDM", 
            "Bulan", "Nominal Cicilan", "Vendor", 
            "No HP", "Status Pembayaran"
        ])

    if "Data Cicilan yang Sudah Dibayar" not in workbook.sheetnames:
        sheet = workbook.create_sheet("Data_Cicilan_yang_Sudah_Dibayar")
        sheet.append([
            "No Kontrak", "Hari dan Tanggal", "Nama BDM", 
            "Bulan", "Nominal Cicilan", "Vendor", 
            "No HP", "Status Pembayaran"
        ])

    if "Ringkasan Data" not in workbook.sheetnames:
        sheet = workbook.create_sheet("Ringkasan_Data")
        sheet.append(["No Kontrak", "Nama BDM", "Sisa Tagihan"])  # Header


    workbook.save(FILE_NAME)

init_sqlite_db()
initialize_excel()


def update_summary_sheet(workbook):
    unpaid_sheet = workbook["Data_Cicilan_yang_Belum_Dibayar"]
    summary_sheet = workbook["Ringkasan_Data"]

    # Hapus semua data sebelumnya, tetapi simpan header
    summary_sheet.delete_rows(2, summary_sheet.max_row)

    # Mengumpulkan data ringkasan
    summary_data = {}
    for row in unpaid_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and isinstance(row[4], (int, float)):  # Pastikan No Kontrak dan Nominal valid
            contract_no = row[0]
            bdm_name = row[2]
            nominal = row[4]

            if contract_no not in summary_data:
                summary_data[contract_no] = {
                    "bdm_name": bdm_name,
                    "sisa_tagihan": 0
                }
            summary_data[contract_no]["sisa_tagihan"] += nominal

    # Menuliskan ringkasan data ke sheet
    for contract_no, data in summary_data.items():
        summary_sheet.append([
            contract_no,
            data["bdm_name"],
            data["sisa_tagihan"]
        ])

    workbook.save(FILE_NAME)



def validate_row(row):
    """
    Validates a row of data to ensure all required fields are present and valid.
    Returns True if valid, False otherwise.
    """
    try:
        if row[0] is None or not isinstance(row[0], (int, str)):  # Contract No
            return False
        if row[1] is not None and not isinstance(row[1], str):  # Date
            return False
        if row[4] is not None and not isinstance(row[4], (int, float)):  # Nominal
            return False
        return True
    except IndexError:
        return False


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = sqlite3.connect("billing_system.db")
        cursor = conn.cursor()

        cursor.execute("SELECT id, username, password, role, approved FROM users WHERE username = ?", (username,))
        user = cursor.fetchone()
        conn.close()

        if user and check_password_hash(user[2], password):
            if user[4] == 1:  # Cek apakah pengguna sudah disetujui
                session.clear()  # Pastikan sesi lama dihapus
                session['user_id'] = user[0]
                session['username'] = user[1]
                session['role'] = user[3]
                flash("Login berhasil!")
                return redirect('/')
            else:
                flash("Akun Anda belum disetujui oleh admin.")
        else:
            flash("Username atau password salah!")
    return render_template("login.html")




@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        nama = request.form['nama'].strip()
        alamat = request.form['alamat'].strip()
        no_hp = request.form['no_hp'].strip()
        email = request.form['email'].strip()

        if not username or not password or not nama or not alamat or not no_hp or not email:
            flash("Semua field harus diisi!")
            return redirect('/register')

        conn = sqlite3.connect("billing_system.db")
        cursor = conn.cursor()
        hashed_password = generate_password_hash(password)

        try:
            cursor.execute('''
                INSERT INTO users (username, password, nama, alamat, no_hp, email, approved) 
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (username, hashed_password, nama, alamat, no_hp, email, 0))  # Default approved = 0
            conn.commit()
            flash("Pendaftaran berhasil! Tunggu persetujuan admin.")
            return redirect('/login')
        except sqlite3.IntegrityError:
            flash("Username sudah terdaftar!")
        finally:
            conn.close()
    return render_template("register.html")



@app.route('/')
def index():
    if 'username' in session:
        return redirect('/form')
    return redirect('/login')

@app.route('/form')
def form():
    if 'username' not in session:
        return redirect('/login')

    conn = sqlite3.connect("billing_system.db")
    cursor = conn.cursor()
    cursor.execute("SELECT nama, email FROM users WHERE username = ?", (session['username'],))
    user_info = cursor.fetchone()
    conn.close()

    return render_template("form.html", user_info=user_info)


@app.route('/submit', methods=["POST"])
def submit():
    if 'username' not in session:
        flash("Anda harus login untuk mengakses fitur ini.")
        return redirect('/login')

    try:
        # Ambil data form
        contract_no = request.form["contract_no"]
        date = request.form["date"]
        bdm_name = request.form["bdm_name"]
        nominal = int(request.form["nominal"])
        duration = int(request.form["duration"])
        vendor_name = request.form.get("vendor_name", "")
        phone = request.form.get("phone", "")

        # Hitung cicilan per bulan
        monthly_payment = nominal // duration

        # Load workbook dan sheet
        workbook = openpyxl.load_workbook(FILE_NAME)
        unpaid_sheet = workbook["Data_Cicilan_yang_Belum_Dibayar"]

        # Ambil nama user
        current_user = session['username']

        # Cari baris terakhir yang digunakan
        last_row = unpaid_sheet.max_row

        # Tambahkan nama user sebagai header
        header_row = last_row + 2
        unpaid_sheet.cell(row=header_row, column=1, value=f"Diinput oleh: {current_user}")
        unpaid_sheet.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=8)

        # Tambahkan data cicilan di bawah header
        for month in range(1, duration + 1):
            unpaid_sheet.append([
                contract_no, date, bdm_name, month, monthly_payment, vendor_name, phone, "Belum Dibayar"
            ])

        # Tambahkan spasi kosong (baris kosong) setelah data
        unpaid_sheet.append(["" for _ in range(8)])
        unpaid_sheet.append(["" for _ in range(8)])


        # Simpan workbook
        update_summary_sheet(workbook)
        workbook.save(FILE_NAME)
        flash("Data berhasil ditambahkan!")

    except Exception as e:
        flash(f"Terjadi kesalahan: {str(e)}")

    return redirect('/report')




@app.route('/report')
def report():
    if 'username' not in session:
        flash("Anda harus login untuk mengakses fitur ini.")
        return redirect('/login')

    try:
        # Load workbook dan sheets
        workbook = openpyxl.load_workbook(FILE_NAME)
        unpaid_sheet = workbook["Data_Cicilan_yang_Belum_Dibayar"]
        paid_sheet = workbook["Data_Cicilan_yang_Sudah_Dibayar"]

        # Proses data cicilan belum dibayar
        unpaid_reports = []
        total_unpaid = 0
        for idx, row in enumerate(unpaid_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not validate_row(row):
                print(f"Invalid row in Unpaid Sheet at line {idx}: {row}")
                continue

            unpaid_reports.append({
                "contract_no": row[0],
                "date": row[1],
                "name": row[2],
                "month": row[3],
                "nominal": row[4],
                "vendor": row[5],
                "phone": row[6],
                "status": row[7],
            })
            total_unpaid += row[4] if row[4] is not None else 0

        # Proses data cicilan yang sudah dibayar
        paid_reports = []
        for idx, row in enumerate(paid_sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Lewati baris yang mengandung "Diinput oleh"
            if row[0] and "Diinput oleh" in row[0]:
                continue

            if not validate_row(row):
                print(f"Invalid row in Paid Sheet at line {idx}: {row}")
                continue

            paid_reports.append({
                "contract_no": row[0],
                "date": row[1],
                "name": row[2],
                "month": row[3],
                "nominal": row[4],
                "vendor": row[5],
                "phone": row[6],
                "status": "Pembayaran Selesai",
            })

        # Ringkasan data
        summary_data = []
        summary_sheet = workbook["Ringkasan_Data"]
        for row in summary_sheet.iter_rows(min_row=2, values_only=True):
            summary_data.append({
                "contract_no": row[0],
                "bdm_name": row[1],
                "sisa_tagihan": row[2],
            })

        # Render halaman laporan
        return render_template(
            "report.html",
            unpaid_reports=unpaid_reports,
            paid_reports=paid_reports,
            total_unpaid=total_unpaid,
            summary_data=summary_data
        )

    except Exception as e:
        flash(f"Terjadi kesalahan saat memuat laporan: {str(e)}")
        return redirect('/')




@app.route('/mark_paid', methods=['POST'])
def mark_paid():
    if 'username' not in session:
        flash("Anda harus login untuk mengakses fitur ini.", "danger")
        return redirect('/login')

    try:
        # Ambil data dari form
        bdm_name = request.form["bdm_name"].strip()
        month_input = request.form.get("month")
        if not month_input or not month_input.isdigit():
            flash("Bulan harus berupa angka valid.", "warning")
            return redirect('/report')

        month = int(month_input)
        current_user = session['username']

        # Load workbook dan sheets
        try:
            workbook = openpyxl.load_workbook(FILE_NAME)
        except PermissionError:
            flash("File sedang digunakan atau tidak bisa diakses. Tutup file Excel terlebih dahulu!", "danger")
            return redirect('/report')

        unpaid_sheet = workbook["Data_Cicilan_yang_Belum_Dibayar"]
        paid_sheet = workbook["Data_Cicilan_yang_Sudah_Dibayar"]

        row_to_delete = None
        data_to_transfer = None

        # Loop menggunakan indeks baris
        for idx, row in enumerate(unpaid_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[2] == bdm_name and row[3] == month:
                data_to_transfer = list(row)
                row_to_delete = idx
                break

        if data_to_transfer and row_to_delete:
            # Tambahkan header jika belum ada untuk user di sheet "Data Cicilan yang Sudah Dibayar"
            user_section_exists = any(
                paid_row[0] and f"Diinput oleh: {current_user}" in paid_row[0]
                for paid_row in paid_sheet.iter_rows(min_row=2, values_only=True)
            )
            if not user_section_exists:
                paid_sheet.append([f"Diinput oleh: {current_user}"] + [""] * 7)

            # Tambahkan data ke sheet "Data Cicilan yang Sudah Dibayar"
            paid_sheet.append(data_to_transfer[:-1] + ["Pembayaran Selesai"])

            # Hapus baris dari "Data Cicilan yang Belum Dibayar"
            unpaid_sheet.delete_rows(row_to_delete)

            # Simpan workbook dengan pengecekan error
            try:
                update_summary_sheet(workbook)
                workbook.save(FILE_NAME)
                flash("Data berhasil dipindahkan ke laporan cicilan yang sudah dibayar.", "success")
            except PermissionError:
                flash("Tidak dapat menyimpan perubahan. File mungkin sedang digunakan.", "danger")
        else:
            flash("Data tidak ditemukan atau sudah dibayar.", "warning")

    except Exception as e:
        flash(f"Terjadi kesalahan: {str(e)}", "danger")

    return redirect('/report')

@app.route('/delete_summary_data', methods=['POST'])
def delete_summary_data():
    if 'username' not in session:
        flash("Anda harus login untuk mengakses fitur ini.")
        return redirect('/login')

    try:
        contract_no = request.form['contract_no']

        workbook = openpyxl.load_workbook(FILE_NAME)
        summary_sheet = workbook["Ringkasan_Data"]

        for idx, row in enumerate(summary_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == contract_no:
                summary_sheet.delete_rows(idx)
                workbook.save(FILE_NAME)
                flash("Data berhasil dihapus!", "success")
                return redirect('/report')

        flash("Data tidak ditemukan!", "warning")
    except Exception as e:
        flash(f"Terjadi kesalahan: {str(e)}", "danger")

    return redirect('/report')

@app.route('/edit_summary_data', methods=["GET", "POST"])
def edit_summary_data():
    if request.method == "GET":
        contract_no = request.args.get("contract_no")
        # Ambil data berdasarkan `contract_no` untuk ditampilkan di form edit
        # Misalnya:
        workbook = openpyxl.load_workbook(FILE_NAME)
        unpaid_sheet = workbook["Data_Cicilan_yang_Belum_Dibayar"]

        # Cari data berdasarkan `contract_no`
        data_to_edit = []
        for row in unpaid_sheet.iter_rows(values_only=True):
            if row[0] == contract_no:  # Asumsi kolom pertama adalah `No Kontrak`
                data_to_edit.append(row)

        if not data_to_edit:
            flash("Data tidak ditemukan.")
            return redirect('/report')

        return render_template("edit_summary.html", data=data_to_edit)

    elif request.method == "POST":
        # Proses update data
        try:
            contract_no = request.form["contract_no"]
            bdm_name = request.form["bdm_name"]
            nominal = int(request.form["nominal"])
            duration = int(request.form["duration"])
            vendor_name = request.form.get("vendor_name", "")
            phone = request.form.get("phone", "")

            # Perbarui data di sheet
            workbook = openpyxl.load_workbook(FILE_NAME)
            unpaid_sheet = workbook["Data_Cicilan_yang_Belum_Dibayar"]

            # Hapus data lama
            rows_to_delete = []
            for row in unpaid_sheet.iter_rows(values_only=False):
                if row[0].value == contract_no:  # Asumsi kolom pertama adalah `No Kontrak`
                    rows_to_delete.append(row[0].row)

            for row_idx in sorted(rows_to_delete, reverse=True):
                unpaid_sheet.delete_rows(row_idx)

            # Tambahkan data baru
            for month in range(1, duration + 1):
                unpaid_sheet.append([
                    contract_no, "", bdm_name, month, nominal // duration, vendor_name, phone, "Belum Dibayar"
                ])

            workbook.save(FILE_NAME)
            flash("Data berhasil diperbarui!")
        except Exception as e:
            flash(f"Terjadi kesalahan: {str(e)}")

        return redirect('/report')

@app.route('/edit/<int:data_id>', methods=["GET"])
def edit(data_id):
    if 'username' not in session:
        flash("Anda harus login untuk mengakses fitur ini.")
        return redirect('/login')

    try:
        # Load workbook dan sheet
        workbook = openpyxl.load_workbook(FILE_NAME)
        unpaid_sheet = workbook["Data_Cicilan_yang_Belum_Dibayar"]

        # Cari data berdasarkan ID
        data = None
        for row in range(1, unpaid_sheet.max_row + 1):
            if unpaid_sheet.cell(row=row, column=9).value == data_id:  # Asumsi ID disimpan di kolom 9
                data = {
                    "id": data_id,
                    "contract_no": unpaid_sheet.cell(row=row, column=1).value,
                    "date": unpaid_sheet.cell(row=row, column=2).value,
                    "bdm_name": unpaid_sheet.cell(row=row, column=3).value,
                    "nominal": unpaid_sheet.cell(row=row, column=5).value * unpaid_sheet.cell(row=row, column=4).value,  # Nominal total
                    "duration": unpaid_sheet.cell(row=row, column=4).value,
                    "vendor_name": unpaid_sheet.cell(row=row, column=6).value,
                    "phone": unpaid_sheet.cell(row=row, column=7).value,
                }
                break

        if not data:
            flash("Data tidak ditemukan.")
            return redirect('/report')

        # Kirim data ke template edit_summary.html
        return render_template("edit_summary.html", data=data)

    except Exception as e:
        flash(f"Terjadi kesalahan: {str(e)}")
        return redirect('/report')


@app.route('/update', methods=["POST"])
def update():
    if 'username' not in session:
        flash("Anda harus login untuk mengakses fitur ini.")
        return redirect('/login')

    try:
        # Ambil data form
        data_id = int(request.form["id"])  # ID baris yang diupdate
        contract_no = request.form["contract_no"]
        date = request.form["date"]
        bdm_name = request.form["bdm_name"]
        nominal = int(request.form["nominal"])
        duration = int(request.form["duration"])
        vendor_name = request.form.get("vendor_name", "")
        phone = request.form.get("phone", "")

        # Hitung cicilan per bulan
        monthly_payment = nominal // duration

        # Load workbook dan sheet
        workbook = openpyxl.load_workbook(FILE_NAME)
        unpaid_sheet = workbook["Data_Cicilan_yang_Belum_Dibayar"]

        # Hapus data lama berdasarkan ID (ID digunakan sebagai penanda baris)
        row_to_delete = None
        for row in range(1, unpaid_sheet.max_row + 1):
            if unpaid_sheet.cell(row=row, column=9).value == data_id:  # Asumsi ID disimpan di kolom 9
                row_to_delete = row
                break
        
        if row_to_delete:
            unpaid_sheet.delete_rows(row_to_delete, 1)  # Hapus baris lama
        else:
            flash("Data tidak ditemukan.")
            return redirect('/report')

        # Tambahkan data baru
        # Cari baris terakhir yang digunakan
        last_row = unpaid_sheet.max_row

        # Tambahkan nama user sebagai header
        current_user = session['username']
        header_row = last_row + 2
        unpaid_sheet.cell(row=header_row, column=1, value=f"Diupdate oleh: {current_user}")
        unpaid_sheet.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=8)

        # Tambahkan data cicilan di bawah header
        for month in range(1, duration + 1):
            unpaid_sheet.append([
                contract_no, date, bdm_name, month, monthly_payment, vendor_name, phone, "Belum Dibayar"
            ])

        # Tambahkan spasi kosong (baris kosong) setelah data
        unpaid_sheet.append(["" for _ in range(8)])
        unpaid_sheet.append(["" for _ in range(8)])

        # Simpan workbook
        update_summary_sheet(workbook)
        workbook.save(FILE_NAME)

        flash("Data berhasil diupdate!")

    except Exception as e:
        flash(f"Terjadi kesalahan: {str(e)}")

    return redirect('/report')


@app.route('/chat', methods=['GET', 'POST'])
def chat():
    if 'username' not in session:
        return redirect('/login')

    conn = sqlite3.connect("billing_system.db")
    cursor = conn.cursor()

    if request.method == 'POST':
        message = request.form['message'].strip()
        if message:
            cursor.execute("INSERT INTO obrolan (username, message) VALUES (?, ?)", (session['username'], message))
            conn.commit()

    cursor.execute("SELECT username, message, timestamp FROM obrolan ORDER BY timestamp ASC")
    chat_history = cursor.fetchall()

    conn.close()
    return render_template("chat.html", chat_history=chat_history)

@app.route('/logout')
def logout():
    session.clear()  # Hapus semua data sesi
    flash("Anda telah logout.")
    return redirect('/login')



@app.route('/admin_requests')
def admin_requests():
    if 'username' not in session or session.get('role') != 'admin':
        flash("Anda tidak memiliki akses ke halaman ini.")
        return redirect('/')

    conn = sqlite3.connect("billing_system.db")
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM users WHERE approved = 0")
    pending_users = cursor.fetchall()
    conn.close()

    return render_template("admin_requests.html", pending_users=pending_users)


@app.route('/approve_user/<int:user_id>', methods=['POST'])
def approve_user(user_id):
    if 'username' not in session or session.get('role') != 'admin':
        flash("Anda tidak memiliki akses untuk tindakan ini.")
        return redirect('/')  # Mengganti redirect('') dengan URL root yang masuk akal

    conn = sqlite3.connect("billing_system.db")
    cursor = conn.cursor()

    # Set kolom approved menjadi 1 untuk user yang sesuai
    cursor.execute("UPDATE users SET approved = 1 WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()

    flash("Pengguna berhasil disetujui.")
    return redirect('/admin_requests')


@app.route('/view_users')
@admin_required
def view_users():
    print("Fungsi view_users dipanggil!")  
    conn = sqlite3.connect("billing_system.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users")
    users = cursor.fetchall()
    conn.close()

    return render_template("view_users.html", users=users)

@app.route('/update_user/<int:user_id>', methods=['GET', 'POST'])
def update_user(user_id):
    if 'username' not in session or session.get('role') != 'admin':
        flash("Anda tidak memiliki akses untuk tindakan ini.")
        return redirect('/')

    conn = sqlite3.connect("billing_system.db")
    cursor = conn.cursor()

    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        role = request.form['role']
        password = request.form.get('password')  # Bisa kosong jika tidak diisi

    if password:  # Jika password baru diisi, hash password tersebut
        hashed_password = generate_password_hash(password)
        cursor.execute("""
            UPDATE users 
            SET username = ?, email = ?, role = ?, password = ?
            WHERE id = ?
        """, (username, email, role, hashed_password, user_id))
    else:  # Jika password tidak diisi, update tanpa mengganti password
        cursor.execute("""
            UPDATE users 
            SET username = ?, email = ?, role = ?
            WHERE id = ?
        """, (username, email, role, user_id))

        
        conn.commit()
        conn.close()

        flash("Data pengguna berhasil diperbarui.")
        return redirect('/view_users')

    # Jika metode GET, ambil data user untuk diisi di form
    cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    user = cursor.fetchone()
    conn.close()

    return render_template("update_user.html", user=user)


@app.route('/delete_user/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    if 'username' not in session or session.get('role') != 'admin':
        flash("Anda tidak memiliki akses untuk tindakan ini.")
        return redirect('/')

    conn = sqlite3.connect("billing_system.db")
    cursor = conn.cursor()

    cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()

    flash("Pengguna berhasil dihapus.")
    return redirect('/view_users')


if __name__ == "__main__":
    app.run(debug=True, use_reloader=False)

app.permanent_session_lifetime = timedelta(minutes=1)

@app.before_request
def check_session():
    if 'username' not in session and request.endpoint != 'login' and request.endpoint not in ['static']:
        if not getattr(request, 'is_redirecting', False):
            flash("Sesi Anda telah berakhir.")
            request.is_redirecting = True
        return redirect('/login')


@app.before_request
def handle_session():
    if 'username' in session:
        session.permanent = False  # Tetapkan sesi sementara
        session.modified = True    # Perbarui waktu sesi
    elif request.endpoint not in ('login', 'register', 'static'):
        flash("Sesi Anda telah berakhir. Silakan login kembali.")
        return redirect('/login')

app.config.update(
    SESSION_COOKIE_SECURE=True,    # Gunakan HTTPS untuk cookie
    SESSION_COOKIE_HTTPONLY=True, # Batasi akses cookie hanya dari HTTP
    SESSION_COOKIE_SAMESITE='Lax' # Hindari CSRF pada domain yang berbeda
)
